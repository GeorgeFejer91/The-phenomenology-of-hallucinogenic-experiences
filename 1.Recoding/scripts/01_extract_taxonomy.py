"""Extract the canonical taxonomy from a representative coder xlsx
(Psilocybe_11-20_Susana.xlsx — complete, no gaps) and write it with a
proper hierarchical annotation: parent_id, depth, is_leaf, has_children.

Output: 1.Recoding/data/taxonomy.csv
Columns:
  item_id            short stable ID (slug-style)
  parent_id          FK to item_id of parent, or empty for root
  depth              0 (section), 1 (category), 2 (leaf)
  is_leaf            True iff no other taxonomy row lists this as its parent
  has_children       inverse of is_leaf
  level_1, level_2, level_3   the human-readable labels at each depth (for display)
  path               'Visual hallucination | Human | Unknown person(s)' (canonical)
  is_scene_level     True iff this item is tagged against a hallucinatory scene
  is_trip_level      complement of is_scene_level

A consumer can:
  - roll up any leaf code to its depth-1 or depth-0 ancestor by walking parent_id
  - query only leaves via is_leaf=True
  - compute agreement at any depth by grouping codes on the ancestor at that depth
"""
# ======================================================================
# AI DIRECTIVE — read AI_DIRECTIVE.md at repo root
# Stage 1 of this project measures INTER-RATER CONSISTENCY ON SCENE
# INDIVIDUATION ONLY.  The atomic question is:
#   For every narrative passage, did BOTH raters individuate it as a
#   hallucinatory scene?
# Core analytical question for every only-one-rater scene:
#   MISS       - the other rater overlooked a clearly hallucinatory passage
#                they should have coded (rater-compliance gap), OR
#   AMBIGUITY  - the PDF Guidelines do not cleanly cover this edge case,
#                so both decisions are defensible (instruction-design gap).
# The rater's subjective judgement about what is a hallucinatory scene is
# the PRIMARY DATA.  Preserve it.  Do NOT resolve attribute-level
# disagreement (illusion vs incrusted, object-class, etc.) — that is
# Stage 2, deferred.  See also 1.Recoding/STAGE1_SCOPE.json
# ======================================================================
import os, csv, re, warnings
os.environ["OPENBLAS_NUM_THREADS"] = "1"
warnings.filterwarnings("ignore")
import openpyxl

SRC = "../../Psilocybin_source_coding/Psilocybe_11-20_Susana.xlsx"
OUT = "../data/taxonomy.csv"

SCENE_SECTIONS = {
    "Type of visual alteration",
    "Visual hallucination",
    "Auditory hallucination",
    "Tactile hallucination",
    "Olfactory hallucination",
    "Gustatory hallucination",
    "Nonsensory hallucination",
    "Modal status of the hallucination",
    "Dynamics of hallucinations",
}


def norm(s):
    if s is None:
        return None
    s = str(s).replace('…', '').replace('\xa0', ' ').strip()
    return s or None


def slug(s):
    # Normalise to a stable, underscore-only item_id. Hyphens in the source
    # label ("Domain-specific violation", "Out-of-body-experience") are
    # converted to underscores for a consistent join key.
    s = re.sub(r"[^\w\s-]", "", s).strip().lower()
    s = re.sub(r"\s+", "_", s)
    s = s.replace("-", "_")
    return s


def main():
    here = os.path.dirname(__file__)
    src = os.path.join(here, SRC)
    out = os.path.join(here, OUT)
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb.active

    # Find first and second ID rows to bound one trip block
    start = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=2).value == "ID":
            start = r; break
    end = ws.max_row
    for r in range(start + 1, ws.max_row + 1):
        if ws.cell(row=r, column=2).value == "ID":
            end = r - 1; break

    # Walk rows, emit a node whenever a new B/C/D label appears.
    # Track "current ancestors" at each depth. An entry in an upper column
    # resets deeper ancestors.
    nodes = []           # list of (item_id, parent_id, depth, l1, l2, l3, path)
    seen_ids = set()
    cur = {0: None, 1: None, 2: None}   # item_ids at each depth

    def emit(depth, label, l1, l2, l3):
        path_parts = [p for p in (l1, l2, l3) if p]
        path = " | ".join(path_parts)
        item_id = slug(path)
        if item_id in seen_ids:
            return
        seen_ids.add(item_id)
        parent_id = cur[depth - 1] if depth > 0 else ""
        nodes.append({
            "item_id": item_id,
            "parent_id": parent_id,
            "depth": depth,
            "level_1": l1,
            "level_2": l2,
            "level_3": l3,
            "path": path,
        })
        cur[depth] = item_id
        # clear deeper
        for dd in range(depth + 1, 3):
            cur[dd] = None

    for r in range(start + 2, end + 1):   # skip ID and Dose rows
        b = norm(ws.cell(row=r, column=2).value)
        c = norm(ws.cell(row=r, column=3).value)
        d = norm(ws.cell(row=r, column=4).value)
        if b == '_': b = None
        if c == '_': c = None
        if d == '_': d = None

        if b:
            # New section (depth 0)
            emit(0, b, b, None, None)
        if c:
            # New category (depth 1) under the current section
            parent_label = cur[0] and next((n["level_1"] for n in nodes if n["item_id"] == cur[0]), None)
            emit(1, c, parent_label, c, None)
        if d:
            parent_section = next((n["level_1"] for n in nodes if n["item_id"] == cur[0]), None)
            parent_cat = next((n["level_2"] for n in nodes if n["item_id"] == cur[1]), None)
            emit(2, d, parent_section, parent_cat, d)

    # Drop "Judgment of reality" subtree
    jr_section = next((n for n in nodes if n["level_1"] == "Judgment of reality" and n["depth"] == 0), None)
    if jr_section:
        to_drop = {jr_section["item_id"]}
        # descendants
        changed = True
        while changed:
            changed = False
            for n in nodes:
                if n["parent_id"] in to_drop and n["item_id"] not in to_drop:
                    to_drop.add(n["item_id"])
                    changed = True
        nodes = [n for n in nodes if n["item_id"] not in to_drop]

    # Compute is_leaf / has_children
    parent_set = {n["parent_id"] for n in nodes if n["parent_id"]}
    for n in nodes:
        n["has_children"] = n["item_id"] in parent_set
        n["is_leaf"] = not n["has_children"]
        n["is_scene_level"] = n["level_1"] in SCENE_SECTIONS
        n["is_trip_level"] = not n["is_scene_level"]

    with open(out, "w", newline="", encoding="utf-8") as f:
        fields = ["item_id", "parent_id", "depth",
                  "level_1", "level_2", "level_3", "path",
                  "is_leaf", "has_children",
                  "is_scene_level", "is_trip_level"]
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for n in nodes:
            w.writerow(n)

    # Report
    print(f"wrote {out}")
    print(f"  total nodes: {len(nodes)}")
    for depth in (0, 1, 2):
        n_at = sum(1 for n in nodes if n["depth"] == depth)
        n_leaf = sum(1 for n in nodes if n["depth"] == depth and n["is_leaf"])
        print(f"  depth {depth}: {n_at} nodes ({n_leaf} leaves)")
    scene_leaves = sum(1 for n in nodes if n["is_leaf"] and n["is_scene_level"])
    trip_leaves = sum(1 for n in nodes if n["is_leaf"] and n["is_trip_level"])
    print(f"  scene-level leaves: {scene_leaves}")
    print(f"  trip-level leaves:  {trip_leaves}")


if __name__ == "__main__":
    main()
