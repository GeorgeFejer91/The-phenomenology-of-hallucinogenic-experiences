"""Build the analysis-ready data frame bundle at
`1.Recoding/analysis/analysis_frame/`.

Output layout:

  analysis_frame/
    README.md
    scenes.csv                  (305 rows, one per canonical scene_id,
                                 with per-trip denominators joined in)
    scenes_x_raters.csv         (~485 rows, one per (scene × rater) pair)
    codes.csv                   (1481 rows, long-format, scene-context joined)
    trip_totals.csv             (40 rows, per-trip aggregates + L1 prevalence)
    taxonomy.csv                (147 rows, pass-through)
    agreement_flags.csv         (664 rows, pass-through)
    rater_style.csv             (80 rows, pass-through)

    stage2_preview/             (attribute-bearing, descriptive only)
      README.md
      attributes_count.csv        (305 × N attr cols, value ∈ {0,1,2})
      attributes_any.csv          (305 × N attr cols, binary any-rater)
      attributes_consensus.csv    (305 × N attr cols, binary both-rater)
      trip_attributes_count.csv   (40 × M trip-level attr cols, 0/1/2)
      attributes_column_map.csv   (colname → taxonomy lookup)

Every CSV is mirrored as `.xlsx` (convenience).
"""
# ======================================================================
# AI DIRECTIVE — read AI_DIRECTIVE.md at repo root
# Stage 1 of this project measures INTER-RATER CONSISTENCY ON SCENE
# INDIVIDUATION ONLY.  The atomic question is:
#   For every narrative passage, did BOTH raters individuate it as a
#   hallucinatory scene?
# Core analytical question for every only-one-rater scene:
#   MISS         - rater-compliance gap, OR
#   GRANULARITY  - lump-vs-split (resolved by data-coding rule), OR
#   AMBIGUITY-2a - under-specified Guideline edge-case, OR
#   AMBIGUITY-2b - phenomenon outside the hallucinatory-scene category.
# Attribute-bearing sheets produced here (stage2_preview/) are
# DESCRIPTIVE ONLY — they present what each rater tagged, without
# adjudicating attribute disagreements.  Stage 2 (attribute reliability)
# is deferred.  See AI_DIRECTIVE.md and 1.Recoding/STAGE1_SCOPE.json.
# ======================================================================
import os, csv, re
from collections import defaultdict, Counter

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

DATA_DIR = "../data"
TRIPS_DIR = "../data/trips"
OUT_DIR = "../analysis/analysis_frame"
STAGE2_DIR = "../analysis/analysis_frame/stage2_preview"


# ---------------- helpers ----------------

def load(name):
    path = os.path.join(os.path.dirname(__file__), DATA_DIR, name)
    with open(path, encoding="utf-8") as f:
        return list(csv.DictReader(f))


def parse_int(s):
    try:
        return int(s)
    except (ValueError, TypeError):
        return None


def truthy(s):
    """Normalise CSV boolean-ish strings to Python bool."""
    return str(s).strip().lower() in ("true", "1", "yes", "t")


def slug(s):
    """Lowercase, non-alphanumeric → '_', collapse repeats, strip edges.
    Guaranteed: result contains no '__' substring, no leading/trailing '_'.
    Empty input → ''.
    """
    if not s:
        return ""
    out = re.sub(r"[^a-z0-9]+", "_", str(s).lower())
    out = re.sub(r"_+", "_", out).strip("_")
    return out


def attr_colname(level_1, level_2, level_3):
    """Column name for an item at the given (L1, L2, L3) path.
    Uses '__' as the level separator so `split('__')` reverses the encoding.
    """
    parts = [slug(level_1), slug(level_2), slug(level_3)]
    parts = [p for p in parts if p]
    return "attr__" + "__".join(parts) if parts else "attr__unknown"


# ---------------- XLSX helpers (reuse pattern from script 17) ----------------

VERDICT_FILL = {
    "CONVERGENT":   "D1FAE5",  # green
    "MISS":         "FEE2E2",  # red
    "GRANULARITY":  "DBEAFE",  # blue
    "AMBIGUITY":    "FEF3C7",  # amber (2a); 2b uses purple below
}
AMB_2B_HEX = "EDE9FE"


def write_csv(path, fieldnames, rows):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)


def write_xlsx(path, fieldnames, rows, *, col_widths=None, verdict_col=None, flavour_col=None):
    """Write an xlsx mirror of a sheet.

    verdict_col: if provided, scene_id + verdict cells get coloured tint
                 based on the verdict value.
    """
    if not HAS_OPENPYXL:
        return
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = os.path.splitext(os.path.basename(path))[0][:31]
    ws.append(fieldnames)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2B2B2B")
    header_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    scene_col_idx = fieldnames.index("scene_id") + 1 if "scene_id" in fieldnames else None
    v_col_idx = fieldnames.index(verdict_col) + 1 if verdict_col and verdict_col in fieldnames else None
    f_col_idx = fieldnames.index(flavour_col) + 1 if flavour_col and flavour_col in fieldnames else None
    for r in rows:
        ws.append([r.get(c, "") for c in fieldnames])
        if v_col_idx:
            row_idx = ws.max_row
            v = r.get(verdict_col, "")
            hex_fill = VERDICT_FILL.get(v)
            if v == "AMBIGUITY" and r.get(flavour_col, "") == "2b":
                hex_fill = AMB_2B_HEX
            if hex_fill:
                fill = PatternFill("solid", fgColor=hex_fill)
                if scene_col_idx:
                    ws.cell(row=row_idx, column=scene_col_idx).fill = fill
                ws.cell(row=row_idx, column=v_col_idx).fill = fill
    if col_widths:
        for i, col in enumerate(fieldnames, start=1):
            letter = ws.cell(row=1, column=i).column_letter
            ws.column_dimensions[letter].width = col_widths.get(col, 18)
    ws.freeze_panes = "B2"
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions
    wrap_al = Alignment(vertical="top", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_al
    wb.save(path)


def emit(name, fieldnames, rows, subdir="", *, col_widths=None, verdict_col=None, flavour_col=None):
    """Emit CSV + XLSX pair under OUT_DIR/subdir/."""
    folder = os.path.join(os.path.dirname(__file__), STAGE2_DIR if subdir == "stage2" else OUT_DIR)
    os.makedirs(folder, exist_ok=True)
    csv_path = os.path.join(folder, f"{name}.csv")
    xlsx_path = os.path.join(folder, f"{name}.xlsx")
    write_csv(csv_path, fieldnames, rows)
    write_xlsx(xlsx_path, fieldnames, rows,
               col_widths=col_widths,
               verdict_col=verdict_col,
               flavour_col=flavour_col)
    return csv_path


# ---------------- main ----------------

def main():
    here = os.path.dirname(__file__)
    os.makedirs(os.path.join(here, OUT_DIR), exist_ok=True)
    os.makedirs(os.path.join(here, STAGE2_DIR), exist_ok=True)

    # ---- Load sources ----
    trips_raw    = load("trips.csv")
    scenes_raw   = load("scenes.csv")
    codes_raw    = load("codes.csv")
    taxonomy     = load("taxonomy.csv")
    agreement    = load("agreement_flags.csv")
    rater_style  = load("rater_style.csv")
    try:
        verdicts_raw = load("stage1_verdicts.csv")
    except FileNotFoundError:
        verdicts_raw = []
    try:
        flat_raw = load("stage1_all_individuations.csv")
    except FileNotFoundError:
        flat_raw = []

    trips_by_id = {t["trip_id"]: t for t in trips_raw}
    verdicts_by_scene = {v["scene_id"]: v for v in verdicts_raw}
    scenes_by_id = {s["scene_id"]: s for s in scenes_raw}
    flat_by_id = {r["scene_id"]: r for r in flat_raw}

    # ---- Per-trip denominators ----
    per_trip_counts = defaultdict(lambda: Counter())
    for s in scenes_raw:
        tid = s["trip_id"]
        status = s["rater_status"]
        per_trip_counts[tid]["n_scenes"] += 1
        per_trip_counts[tid][f"n_{status}"] += 1  # n_both, n_only_A, n_only_B
        # Derive verdict counts
        if status == "both":
            per_trip_counts[tid]["n_convergent"] += 1
        else:
            per_trip_counts[tid]["n_solo"] += 1
            v = verdicts_by_scene.get(s["scene_id"], {})
            verdict = v.get("verdict", "")
            flavour = v.get("verdict_flavour", "")
            if verdict == "MISS":
                per_trip_counts[tid]["n_miss"] += 1
            elif verdict == "GRANULARITY":
                per_trip_counts[tid]["n_granularity"] += 1
            elif verdict == "AMBIGUITY":
                if flavour == "2a":
                    per_trip_counts[tid]["n_ambiguity_2a"] += 1
                elif flavour == "2b":
                    per_trip_counts[tid]["n_ambiguity_2b"] += 1

    # rater-A / rater-B individuated totals per trip
    for tid, ctr in per_trip_counts.items():
        n_both = ctr.get("n_both", 0)
        ctr["n_scenes_rater_A"] = n_both + ctr.get("n_only_A", 0)
        ctr["n_scenes_rater_B"] = n_both + ctr.get("n_only_B", 0)

    # ---- Scene rank-in-trip ----
    scene_rank = {}
    by_trip_scenes = defaultdict(list)
    for s in scenes_raw:
        by_trip_scenes[s["trip_id"]].append(s)
    for tid, ss in by_trip_scenes.items():
        ss_sorted = sorted(ss, key=lambda s: parse_int(s["canonical_span_start"]) or 0)
        for rank, s in enumerate(ss_sorted, start=1):
            scene_rank[s["scene_id"]] = rank

    # ---- Codes grouped by scene_id and by trip_id ----
    codes_by_scene_item_rater = defaultdict(lambda: Counter())  # (scene_id, item_id) -> {A, B}
    codes_by_trip_item_rater = defaultdict(lambda: Counter())   # (trip_id, item_id) -> {A, B}
    codes_per_rater_per_trip_scene = defaultdict(int)           # (trip_id, rater, scene_level_bool) -> count
    codes_per_rater_per_scene = defaultdict(int)                # (scene_id, rater, scene_level_bool) -> count

    for c in codes_raw:
        sid   = c.get("scene_id") or ""
        tid   = c["trip_id"]
        rater = c["rater"]  # "A" | "B"
        item  = c["item_id"]
        is_sl = truthy(c.get("is_scene_level", "False"))
        if is_sl and sid:
            codes_by_scene_item_rater[(sid, item)][rater] += 1
            codes_per_rater_per_scene[(sid, rater, True)] += 1
            codes_per_rater_per_trip_scene[(tid, rater, True)] += 1
        elif not is_sl:
            codes_by_trip_item_rater[(tid, item)][rater] += 1
            codes_per_rater_per_trip_scene[(tid, rater, False)] += 1

    # ---- Taxonomy lookup ----
    taxonomy_by_id = {t["item_id"]: t for t in taxonomy}

    # ---- Attribute column names + column map ----
    # One column per scene-level item the taxonomy knows about
    scene_level_items = [t for t in taxonomy if truthy(t["is_scene_level"])]
    trip_level_items  = [t for t in taxonomy if truthy(t["is_trip_level"])]

    def build_colmap(items, carried_context_cols):
        """For each taxonomy item return a unique column name.
        On collision, append the item_id to the losing column.
        """
        seen = {}  # colname -> item_id
        rows = []
        for t in items:
            base = attr_colname(t["level_1"], t["level_2"], t["level_3"])
            col = base
            if col in seen and seen[col] != t["item_id"]:
                # collision — disambiguate with item_id slug
                col = base + "__id_" + slug(t["item_id"])
            seen[col] = t["item_id"]
            rows.append({
                "column_name": col,
                "item_id": t["item_id"],
                "level_1": t["level_1"],
                "level_2": t["level_2"],
                "level_3": t["level_3"],
                "depth": t["depth"],
                "is_leaf": t["is_leaf"],
                "is_scene_level": t["is_scene_level"],
                "is_trip_level": t["is_trip_level"],
            })
        return rows

    scene_column_map = build_colmap(scene_level_items, [])
    trip_column_map  = build_colmap(trip_level_items, [])

    # Restrict to items that actually appear in codes.csv (skip zero-sum columns)
    scene_items_used = {c["item_id"] for c in codes_raw if truthy(c.get("is_scene_level")) and c.get("scene_id")}
    trip_items_used  = {c["item_id"] for c in codes_raw if not truthy(c.get("is_scene_level"))}
    scene_column_map = [r for r in scene_column_map if r["item_id"] in scene_items_used]
    trip_column_map  = [r for r in trip_column_map  if r["item_id"] in trip_items_used]

    # Emit combined column map (scene + trip, with scope marker)
    combined_colmap = []
    for r in scene_column_map:
        combined_colmap.append({**r, "scope": "scene"})
    for r in trip_column_map:
        combined_colmap.append({**r, "scope": "trip"})

    scene_colname_by_item = {r["item_id"]: r["column_name"] for r in scene_column_map}
    trip_colname_by_item  = {r["item_id"]: r["column_name"] for r in trip_column_map}

    # ---- Build scenes.csv (enriched) ----
    # Load narratives to compute scene_word_count in characters
    def narrative_of(tid):
        p = os.path.join(here, TRIPS_DIR, f"{tid}.txt")
        return open(p, encoding="utf-8").read() if os.path.exists(p) else ""

    narratives = {tid: narrative_of(tid) for tid in trips_by_id}

    def verdict_class_for_scene(s):
        """Return (verdict, flavour) — CONVERGENT for shared, else from verdicts."""
        if s["rater_status"] == "both":
            return ("CONVERGENT", "")
        v = verdicts_by_scene.get(s["scene_id"], {})
        return (v.get("verdict", ""), v.get("verdict_flavour", ""))

    SCENES_COLS = [
        "scene_id", "trip_id", "substance", "block", "coder_A", "coder_B",
        "rater_status", "rater_A_individuated", "rater_B_individuated",
        "stage1_driver", "parent_scene_id",
        "stage1_verdict", "verdict_flavour", "verdict_source",
        "canonical_span_start", "canonical_span_end",
        "scene_word_count", "scene_rank_in_trip",
        "canonical_desc", "rater_A_excerpt", "rater_B_excerpt",
        "narrative_passage", "stage1_rationale", "verdict_rationale",
        # --- denominators ---
        "n_scenes_in_trip", "n_convergent_in_trip",
        "n_solo_A_in_trip", "n_solo_B_in_trip",
        "n_scenes_rater_A_in_trip", "n_scenes_rater_B_in_trip",
        "n_miss_in_trip", "n_granularity_in_trip",
        "n_ambiguity_2a_in_trip", "n_ambiguity_2b_in_trip",
        "trip_word_count",
    ]

    scenes_rows = []
    for s in scenes_raw:
        tid = s["trip_id"]
        trip = trips_by_id.get(tid, {})
        ctr = per_trip_counts.get(tid, Counter())
        narr = narratives.get(tid, "")
        a = parse_int(s.get("canonical_span_start"))
        b = parse_int(s.get("canonical_span_end"))
        passage = ""
        swc = 0
        if a is not None and b is not None and 0 <= a < len(narr):
            passage = narr[a:min(b, len(narr))]
            swc = max(0, b - a)
        status = s["rater_status"]
        verdict, flavour = verdict_class_for_scene(s)
        vrow = verdicts_by_scene.get(s["scene_id"], {})
        scenes_rows.append({
            "scene_id": s["scene_id"],
            "trip_id": tid,
            "substance": trip.get("substance", ""),
            "block": trip.get("block", ""),
            "coder_A": trip.get("coder_A", ""),
            "coder_B": trip.get("coder_B", ""),
            "rater_status": status,
            "rater_A_individuated": status in ("both", "only_A"),
            "rater_B_individuated": status in ("both", "only_B"),
            "stage1_driver": s.get("stage1_driver", "") or "",
            "parent_scene_id": (vrow.get("parent_scene_id") or s.get("parent_scene_id") or ""),
            "stage1_verdict": verdict,
            "verdict_flavour": flavour,
            "verdict_source": vrow.get("verdict_source", "") or ("auto" if verdict == "CONVERGENT" else ""),
            "canonical_span_start": a if a is not None else "",
            "canonical_span_end":   b if b is not None else "",
            "scene_word_count": swc,
            "scene_rank_in_trip": scene_rank.get(s["scene_id"], ""),
            "canonical_desc": s.get("canonical_desc", "") or "",
            "rater_A_excerpt": s.get("rater_A_refs", "") or "",
            "rater_B_excerpt": s.get("rater_B_refs", "") or "",
            "narrative_passage": passage,
            "stage1_rationale": s.get("stage1_rationale", "") or "",
            "verdict_rationale": vrow.get("rationale", "") or (
                "Both raters individuated this passage." if verdict == "CONVERGENT" else ""),
            # denominators
            "n_scenes_in_trip":         ctr.get("n_scenes", 0),
            "n_convergent_in_trip":     ctr.get("n_convergent", 0),
            "n_solo_A_in_trip":         ctr.get("n_only_A", 0),
            "n_solo_B_in_trip":         ctr.get("n_only_B", 0),
            "n_scenes_rater_A_in_trip": ctr.get("n_scenes_rater_A", 0),
            "n_scenes_rater_B_in_trip": ctr.get("n_scenes_rater_B", 0),
            "n_miss_in_trip":           ctr.get("n_miss", 0),
            "n_granularity_in_trip":    ctr.get("n_granularity", 0),
            "n_ambiguity_2a_in_trip":   ctr.get("n_ambiguity_2a", 0),
            "n_ambiguity_2b_in_trip":   ctr.get("n_ambiguity_2b", 0),
            "trip_word_count": parse_int(trip.get("word_count", "")) or 0,
        })

    # Sort: substance (brug first), trip number, scene rank
    scenes_rows.sort(key=lambda r: (
        0 if r["substance"] == "brugmansia" else 1,
        int(r["trip_id"].split("_")[-1]),
        r["scene_rank_in_trip"] or 0,
    ))

    scene_col_widths = {
        "scene_id": 32, "trip_id": 16, "substance": 12, "block": 8,
        "coder_A": 14, "coder_B": 14, "rater_status": 10,
        "rater_A_individuated": 10, "rater_B_individuated": 10,
        "stage1_driver": 8, "parent_scene_id": 28,
        "stage1_verdict": 14, "verdict_flavour": 8, "verdict_source": 10,
        "canonical_span_start": 10, "canonical_span_end": 10,
        "scene_word_count": 10, "scene_rank_in_trip": 10,
        "canonical_desc": 55, "rater_A_excerpt": 30, "rater_B_excerpt": 30,
        "narrative_passage": 80, "stage1_rationale": 45, "verdict_rationale": 80,
        "n_scenes_in_trip": 10, "n_convergent_in_trip": 12,
        "n_solo_A_in_trip": 10, "n_solo_B_in_trip": 10,
        "n_scenes_rater_A_in_trip": 12, "n_scenes_rater_B_in_trip": 12,
        "n_miss_in_trip": 10, "n_granularity_in_trip": 12,
        "n_ambiguity_2a_in_trip": 12, "n_ambiguity_2b_in_trip": 12,
        "trip_word_count": 10,
    }
    emit("scenes", SCENES_COLS, scenes_rows,
         col_widths=scene_col_widths,
         verdict_col="stage1_verdict", flavour_col="verdict_flavour")

    # ---- scenes_x_raters.csv (one row per (scene × rater)) ----
    SXR_COLS = [
        "scene_id", "trip_id", "substance", "block", "rater", "coder_name",
        "rater_role", "stage1_verdict", "verdict_flavour",
        "scene_excerpt_by_this_rater",
        "n_scene_codes_by_this_rater", "n_trip_codes_by_this_rater",
        "n_scenes_rater_indiv_in_trip", "n_scenes_in_trip",
    ]

    sxr_rows = []
    for s in scenes_raw:
        tid = s["trip_id"]
        trip = trips_by_id.get(tid, {})
        ctr = per_trip_counts.get(tid, Counter())
        verdict, flavour = verdict_class_for_scene(s)
        if s["rater_status"] in ("both", "only_A"):
            sxr_rows.append({
                "scene_id": s["scene_id"],
                "trip_id": tid,
                "substance": trip.get("substance", ""),
                "block": trip.get("block", ""),
                "rater": "A",
                "coder_name": trip.get("coder_A", ""),
                "rater_role": "convergent" if s["rater_status"] == "both" else "solo_A",
                "stage1_verdict": verdict,
                "verdict_flavour": flavour,
                "scene_excerpt_by_this_rater": s.get("rater_A_refs", "") or "",
                "n_scene_codes_by_this_rater": codes_per_rater_per_scene.get((s["scene_id"], "A", True), 0),
                "n_trip_codes_by_this_rater": codes_per_rater_per_trip_scene.get((tid, "A", False), 0),
                "n_scenes_rater_indiv_in_trip": ctr.get("n_scenes_rater_A", 0),
                "n_scenes_in_trip": ctr.get("n_scenes", 0),
            })
        if s["rater_status"] in ("both", "only_B"):
            sxr_rows.append({
                "scene_id": s["scene_id"],
                "trip_id": tid,
                "substance": trip.get("substance", ""),
                "block": trip.get("block", ""),
                "rater": "B",
                "coder_name": trip.get("coder_B", ""),
                "rater_role": "convergent" if s["rater_status"] == "both" else "solo_B",
                "stage1_verdict": verdict,
                "verdict_flavour": flavour,
                "scene_excerpt_by_this_rater": s.get("rater_B_refs", "") or "",
                "n_scene_codes_by_this_rater": codes_per_rater_per_scene.get((s["scene_id"], "B", True), 0),
                "n_trip_codes_by_this_rater": codes_per_rater_per_trip_scene.get((tid, "B", False), 0),
                "n_scenes_rater_indiv_in_trip": ctr.get("n_scenes_rater_B", 0),
                "n_scenes_in_trip": ctr.get("n_scenes", 0),
            })

    sxr_rows.sort(key=lambda r: (
        0 if r["substance"] == "brugmansia" else 1,
        int(r["trip_id"].split("_")[-1]),
        r["scene_id"], r["rater"],
    ))
    sxr_widths = {
        "scene_id": 32, "trip_id": 16, "substance": 12, "block": 8,
        "rater": 6, "coder_name": 14, "rater_role": 12,
        "stage1_verdict": 14, "verdict_flavour": 8,
        "scene_excerpt_by_this_rater": 30,
        "n_scene_codes_by_this_rater": 12, "n_trip_codes_by_this_rater": 12,
        "n_scenes_rater_indiv_in_trip": 12, "n_scenes_in_trip": 12,
    }
    emit("scenes_x_raters", SXR_COLS, sxr_rows,
         col_widths=sxr_widths,
         verdict_col="stage1_verdict", flavour_col="verdict_flavour")

    # ---- codes.csv (enriched with scene-context) ----
    CODES_COLS = list(codes_raw[0].keys()) + [
        "scene_rater_status", "scene_verdict", "scene_verdict_flavour",
        "scene_canonical_desc", "scene_span_start", "scene_span_end",
    ]
    codes_rows = []
    for c in codes_raw:
        enriched = dict(c)
        sid = c.get("scene_id") or ""
        if sid and sid in scenes_by_id:
            s = scenes_by_id[sid]
            verdict, flavour = verdict_class_for_scene(s)
            enriched.update({
                "scene_rater_status": s["rater_status"],
                "scene_verdict": verdict,
                "scene_verdict_flavour": flavour,
                "scene_canonical_desc": s.get("canonical_desc", "") or "",
                "scene_span_start": s.get("canonical_span_start", "") or "",
                "scene_span_end":   s.get("canonical_span_end", "") or "",
            })
        else:
            enriched.update({
                "scene_rater_status": "",
                "scene_verdict": "",
                "scene_verdict_flavour": "",
                "scene_canonical_desc": "",
                "scene_span_start": "",
                "scene_span_end": "",
            })
        codes_rows.append(enriched)
    emit("codes", CODES_COLS, codes_rows,
         col_widths={"excerpt": 50, "scene_canonical_desc": 40, "item_path": 35, "item_id": 30},
         verdict_col="scene_verdict", flavour_col="scene_verdict_flavour")

    # ---- trip_totals.csv ----
    # per-(trip, L1-section) prevalence: how many scenes carry an item under that L1 (any rater / both raters)
    L1_SECTIONS = sorted({t["level_1"] for t in scene_level_items if t["level_1"]})
    L1_any   = defaultdict(lambda: defaultdict(set))  # tid -> L1 -> set(scene_id)
    L1_both  = defaultdict(lambda: defaultdict(set))
    for (sid, item), raters in codes_by_scene_item_rater.items():
        tax = taxonomy_by_id.get(item, {})
        L1 = tax.get("level_1", "")
        if not L1:
            continue
        tid = scenes_by_id[sid]["trip_id"] if sid in scenes_by_id else ""
        if not tid:
            continue
        L1_any[tid][L1].add(sid)
        if raters.get("A", 0) >= 1 and raters.get("B", 0) >= 1:
            L1_both[tid][L1].add(sid)

    TRIP_COLS = [
        "trip_id", "substance", "block", "coder_A", "coder_B", "word_count",
        "n_scenes", "n_convergent", "n_solo_A", "n_solo_B",
        "n_scenes_rater_A", "n_scenes_rater_B",
        "n_miss", "n_granularity", "n_ambiguity_2a", "n_ambiguity_2b",
    ]
    # Append L1-prevalence columns
    for L1 in L1_SECTIONS:
        TRIP_COLS.append(f"n_scenes_with_L1_any__{slug(L1)}")
        TRIP_COLS.append(f"n_scenes_with_L1_consensus__{slug(L1)}")

    trips_rows = []
    for tid, trip in sorted(trips_by_id.items(),
                            key=lambda kv: (0 if kv[1]["substance"] == "brugmansia" else 1,
                                            int(kv[0].split("_")[-1]))):
        ctr = per_trip_counts.get(tid, Counter())
        row = {
            "trip_id": tid,
            "substance": trip.get("substance", ""),
            "block": trip.get("block", ""),
            "coder_A": trip.get("coder_A", ""),
            "coder_B": trip.get("coder_B", ""),
            "word_count": parse_int(trip.get("word_count", "")) or 0,
            "n_scenes": ctr.get("n_scenes", 0),
            "n_convergent": ctr.get("n_convergent", 0),
            "n_solo_A": ctr.get("n_only_A", 0),
            "n_solo_B": ctr.get("n_only_B", 0),
            "n_scenes_rater_A": ctr.get("n_scenes_rater_A", 0),
            "n_scenes_rater_B": ctr.get("n_scenes_rater_B", 0),
            "n_miss": ctr.get("n_miss", 0),
            "n_granularity": ctr.get("n_granularity", 0),
            "n_ambiguity_2a": ctr.get("n_ambiguity_2a", 0),
            "n_ambiguity_2b": ctr.get("n_ambiguity_2b", 0),
        }
        for L1 in L1_SECTIONS:
            row[f"n_scenes_with_L1_any__{slug(L1)}"] = len(L1_any[tid][L1])
            row[f"n_scenes_with_L1_consensus__{slug(L1)}"] = len(L1_both[tid][L1])
        trips_rows.append(row)
    emit("trip_totals", TRIP_COLS, trips_rows,
         col_widths={c: 10 for c in TRIP_COLS if c.startswith("n_")})

    # ---- pass-through sheets ----
    emit("taxonomy", list(taxonomy[0].keys()), taxonomy,
         col_widths={"item_id": 40, "level_1": 22, "level_2": 22, "level_3": 25, "path": 50})
    emit("agreement_flags", list(agreement[0].keys()), agreement,
         col_widths={"scene_id": 32, "item_id": 40, "level_1": 22, "level_2": 22, "level_3": 25, "item_path": 50})
    emit("rater_style", list(rater_style[0].keys()), rater_style,
         col_widths={"trip_id": 16, "coder_name": 14})

    # ---- stage2_preview/ attribute matrices ----
    scene_attr_cols = [r["column_name"] for r in scene_column_map]
    scene_attr_cols_sorted = sorted(scene_attr_cols)

    # attributes_count: 0/1/2 per (scene, item)
    attr_count_rows = []
    attr_any_rows = []
    attr_consensus_rows = []
    for s in scenes_rows:  # sorted already
        sid = s["scene_id"]
        base = {
            "scene_id": sid,
            "trip_id":  s["trip_id"],
            "substance": s["substance"],
            "block":    s["block"],
            "rater_status": s["rater_status"],
            "stage1_verdict": s["stage1_verdict"],
            "verdict_flavour": s["verdict_flavour"],
        }
        count_row = dict(base)
        any_row = dict(base)
        cons_row = dict(base)
        for r in scene_column_map:
            col = r["column_name"]; item = r["item_id"]
            raters = codes_by_scene_item_rater.get((sid, item), Counter())
            n_raters = sum(1 for rr in ("A", "B") if raters.get(rr, 0) > 0)
            count_row[col] = n_raters
            any_row[col]   = 1 if n_raters >= 1 else 0
            cons_row[col]  = 1 if n_raters >= 2 else 0
        attr_count_rows.append(count_row)
        attr_any_rows.append(any_row)
        attr_consensus_rows.append(cons_row)

    ATTR_BASE_COLS = ["scene_id", "trip_id", "substance", "block", "rater_status",
                      "stage1_verdict", "verdict_flavour"]
    ATTR_COLS = ATTR_BASE_COLS + scene_attr_cols_sorted
    attr_widths = {c: 18 for c in ATTR_BASE_COLS}
    attr_widths["scene_id"] = 32
    emit("attributes_count", ATTR_COLS, attr_count_rows, subdir="stage2",
         col_widths=attr_widths,
         verdict_col="stage1_verdict", flavour_col="verdict_flavour")
    emit("attributes_any", ATTR_COLS, attr_any_rows, subdir="stage2",
         col_widths=attr_widths,
         verdict_col="stage1_verdict", flavour_col="verdict_flavour")
    emit("attributes_consensus", ATTR_COLS, attr_consensus_rows, subdir="stage2",
         col_widths=attr_widths,
         verdict_col="stage1_verdict", flavour_col="verdict_flavour")

    # trip_attributes_count.csv
    trip_attr_cols = [r["column_name"] for r in trip_column_map]
    trip_attr_cols_sorted = sorted(trip_attr_cols)
    trip_attr_rows = []
    for trow in trips_rows:
        tid = trow["trip_id"]
        base = {
            "trip_id": tid, "substance": trow["substance"], "block": trow["block"],
            "coder_A": trow["coder_A"], "coder_B": trow["coder_B"],
            "word_count": trow["word_count"], "n_scenes": trow["n_scenes"],
        }
        for r in trip_column_map:
            col = r["column_name"]; item = r["item_id"]
            raters = codes_by_trip_item_rater.get((tid, item), Counter())
            n_raters = sum(1 for rr in ("A", "B") if raters.get(rr, 0) > 0)
            base[col] = n_raters
        trip_attr_rows.append(base)
    TRIP_ATTR_COLS = ["trip_id", "substance", "block", "coder_A", "coder_B",
                      "word_count", "n_scenes"] + trip_attr_cols_sorted
    emit("trip_attributes_count", TRIP_ATTR_COLS, trip_attr_rows, subdir="stage2",
         col_widths={"trip_id": 16, "substance": 12, "coder_A": 14, "coder_B": 14})

    # attributes_column_map.csv (combined scene + trip)
    COLMAP_COLS = ["column_name", "scope", "item_id", "level_1", "level_2", "level_3",
                   "depth", "is_leaf", "is_scene_level", "is_trip_level"]
    emit("attributes_column_map", COLMAP_COLS, combined_colmap, subdir="stage2",
         col_widths={"column_name": 60, "item_id": 45, "level_1": 22, "level_2": 22, "level_3": 25})

    # ---- README (top level) ----
    readme_top = f"""# Analysis-ready data frame bundle

*Stage 1 of the phenomenology-of-hallucinations pipeline — see
`../../AI_DIRECTIVE.md` at the repo root.*

This bundle re-shapes the canonical pipeline outputs (`1.Recoding/data/*.csv`)
into a single analytic scaffold where `scene_id` is the primary cross-reference
key, per-trip denominators are pre-computed, and the taxonomy hierarchy is
exposed for nested-level queries.

## Files (top level — Stage-1-clean)

| file | grain | rows | notes |
|---|---|---|---|
| `scenes.csv`/`.xlsx` | per canonical scene_id | {len(scenes_rows)} | all metadata + Stage-1 verdict + per-trip denominators |
| `scenes_x_raters.csv`/`.xlsx` | per (scene × rater who individuated it) | {len(sxr_rows)} | natural grain for rater-effect models |
| `codes.csv`/`.xlsx` | per (scene × rater × item) coding event | {len(codes_rows)} | long-format with scene-context joined |
| `trip_totals.csv`/`.xlsx` | per trip | {len(trips_rows)} | scene-count + verdict-count + per-L1 prevalence |
| `taxonomy.csv`/`.xlsx` | per taxonomy item | {len(taxonomy)} | pass-through convenience copy |
| `agreement_flags.csv`/`.xlsx` | per (scene × item) | {len(agreement)} | AGREE / A_ONLY / B_ONLY at any taxonomy depth |
| `rater_style.csv`/`.xlsx` | per (trip × rater) | {len(rater_style)} | per-rater tagging density / leaf preferences |

**CSV is canonical.** XLSX is a convenience mirror with coloured verdict
cells, freeze-panes and auto-filter. If the two ever disagree, trust
the CSV.

## Stage-1 / Stage-2 boundary

This top-level bundle only exposes **scene individuation** state (Stage 1).
It reports each rater's tags but never adjudicates attribute disagreements.

Attribute-bearing wide matrices live separately under `stage2_preview/` and
are **descriptive only** — they present what each rater tagged without
implying that any single coding is "correct". Stage 2 (attribute-level
reliability) is deferred per `AI_DIRECTIVE.md`.

## Key columns on `scenes.csv`

- `scene_id`: unique code. Shared scenes end in `_AB` (one code, both
  raters individuated). Solo scenes end in rater-specific driver
  suffixes (`_A_RCL`, `_B_AMP`, etc.).
- `rater_status` ∈ `{{both, only_A, only_B}}`.
- `stage1_verdict` ∈ `{{CONVERGENT, MISS, GRANULARITY, AMBIGUITY}}`, with
  `verdict_flavour` ∈ `{{2a, 2b}}` for AMBIGUITY.
- `n_scenes_in_trip`, `n_convergent_in_trip`, `n_scenes_rater_A_in_trip`,
  `n_scenes_rater_B_in_trip`: denominators for proportion analyses.
- `scene_word_count`, `trip_word_count`: for per-1000-words normalisation.
- `parent_scene_id`: for GRANULARITY fragments, points to the larger
  canonical parent scene under which the fragment pools for downstream
  attribute analysis.

## Example queries

### A. "Proportion of individuated scenes (any rater) that include an animal-class hallucination, per substance, normalised per trip"

1. Start from `scenes.csv` joined with `stage2_preview/attributes_any.csv` on `scene_id`.
2. Sum any column matching `attr__visual_hallucination__animal*` per scene → a `has_animal` 0/1 flag.
3. Group by `trip_id`: `n_animal_scenes / n_scenes_in_trip` → per-trip rate.
4. Group by `substance`: take the mean (or median, or full distribution) of per-trip rates.

```python
import pandas as pd, re
scenes = pd.read_csv("scenes.csv")
attrs  = pd.read_csv("stage2_preview/attributes_any.csv")
animal_cols = [c for c in attrs.columns if c.startswith("attr__visual_hallucination__animal")]
attrs["has_animal"] = (attrs[animal_cols].sum(axis=1) > 0).astype(int)
df = scenes.merge(attrs[["scene_id", "has_animal"]], on="scene_id")
per_trip = df.groupby(["substance", "trip_id"]).agg(
    n_scenes=("scene_id", "count"),
    n_animal=("has_animal", "sum"),
).reset_index()
per_trip["rate"] = per_trip["n_animal"] / per_trip["n_scenes"]
per_trip.groupby("substance")["rate"].describe()
```

### B. "Same, but restricted to convergent scenes (raters agree on scene existence)"

Add `df = df[df["rater_status"] == "both"]` before the group-by, and use
`n_convergent_in_trip` as the denominator.

### C. "Same, but restricted to convergent scenes WHERE BOTH RATERS TAGGED IT AS ANIMAL"

Replace `attributes_any.csv` with `attributes_consensus.csv`. This filters
attribute agreement on top of scene-existence agreement.

### D. "Rate per 1000 words of narrative"

Divide `n_animal` by `trip_word_count / 1000` instead of scene count.

## Taxonomy hierarchy (nested analyses)

Every attribute column on the wide matrices is named
`attr__{{L1_slug}}__{{L2_slug}}__{{L3_slug}}`. Missing levels are omitted;
double-underscore is the separator; single-underscore only appears inside
a slug. To sum at any level, grep by prefix:

| cut | regex |
|---|---|
| all Visual-hallucination items | `^attr__visual_hallucination__` |
| all Visual-hallucination Animals | `^attr__visual_hallucination__animal__` |
| all Visual-hallucination Animal Insects | `^attr__visual_hallucination__animal__insect$` |

`stage2_preview/attributes_column_map.csv` gives the full lookup
(column_name → item_id, level_1/2/3, depth, is_leaf) for programmatic use.

## Regenerating the bundle

From `1.Recoding/scripts/`:

```sh
py 18_build_analysis_frame.py
```

The script is idempotent and purely additive — it never modifies the
canonical `data/*.csv` files.
"""
    with open(os.path.join(here, OUT_DIR, "README.md"), "w", encoding="utf-8") as f:
        f.write(readme_top)

    # ---- README (stage2_preview) ----
    readme_s2 = f"""# stage2_preview — attribute-bearing artefacts (DESCRIPTIVE ONLY)

**Read `AI_DIRECTIVE.md` at the repo root first.**

Everything in this folder is a convenience view over each rater's
attribute tags. Stage 1 measures *scene-individuation* reliability only;
attribute-level reliability (Stage 2) is explicitly deferred. These
files report what each rater tagged without adjudicating.

## Files

| file | value semantics | rows × cols | notes |
|---|---|---|---|
| `attributes_count.csv` / `.xlsx` | number of raters (0/1/2) who tagged this scene with this item | {len(scenes_rows)} × {len(ATTR_COLS)} | primary, information-preserving |
| `attributes_any.csv` / `.xlsx` | 1 if ≥1 rater tagged it, else 0 | {len(scenes_rows)} × {len(ATTR_COLS)} | convenience (≥1 rater) |
| `attributes_consensus.csv` / `.xlsx` | 1 if BOTH raters tagged it, else 0 | {len(scenes_rows)} × {len(ATTR_COLS)} | convenience (consensus filter) |
| `trip_attributes_count.csv` / `.xlsx` | 0/1/2 for trip-level items | {len(trip_attr_rows)} × {len(TRIP_ATTR_COLS)} | separate grain — never duplicate into scene sheets |
| `attributes_column_map.csv` | column_name → taxonomy lookup | {len(combined_colmap)} × {len(COLMAP_COLS)} | use this when parsing column names programmatically |

## Column naming

`attr__{{L1_slug}}__{{L2_slug}}__{{L3_slug}}`.
- Double underscore separates levels.
- Single underscore only appears inside a slug.
- Parsing: `colname.removeprefix("attr__").split("__")`.

## Carried-context columns on every wide sheet

`scene_id, trip_id, substance, block, rater_status, stage1_verdict,
verdict_flavour` — so substance-level proportions don't require joining
back to `scenes.csv`.

## What these sheets DO NOT give you

- Reliability of attribute choices on a given scene. Two raters disagreeing
  about illusion-vs-incrusted-object is recorded in `attributes_count` as
  each having tagged their own preferred leaf; neither choice is endorsed.
- Consensus on modal-status / dynamics / object-class. Those require
  Stage-2 adjudication (out of scope).
"""
    with open(os.path.join(here, STAGE2_DIR, "README.md"), "w", encoding="utf-8") as f:
        f.write(readme_s2)

    # ---- summary ----
    print()
    print("=" * 72)
    print("Analysis frame written to:", os.path.abspath(os.path.join(here, OUT_DIR)))
    print("=" * 72)
    print(f"  scenes.csv                    {len(scenes_rows)} rows × {len(SCENES_COLS)} cols")
    print(f"  scenes_x_raters.csv           {len(sxr_rows)} rows × {len(SXR_COLS)} cols")
    print(f"  codes.csv                     {len(codes_rows)} rows × {len(CODES_COLS)} cols")
    print(f"  trip_totals.csv               {len(trips_rows)} rows × {len(TRIP_COLS)} cols")
    print(f"  taxonomy.csv                  {len(taxonomy)} rows")
    print(f"  agreement_flags.csv           {len(agreement)} rows")
    print(f"  rater_style.csv               {len(rater_style)} rows")
    print()
    print("  stage2_preview/")
    print(f"    attributes_count.csv        {len(attr_count_rows)} rows × {len(ATTR_COLS)} cols")
    print(f"    attributes_any.csv          {len(attr_any_rows)} rows × {len(ATTR_COLS)} cols")
    print(f"    attributes_consensus.csv    {len(attr_consensus_rows)} rows × {len(ATTR_COLS)} cols")
    print(f"    trip_attributes_count.csv   {len(trip_attr_rows)} rows × {len(TRIP_ATTR_COLS)} cols")
    print(f"    attributes_column_map.csv   {len(combined_colmap)} rows × {len(COLMAP_COLS)} cols")


if __name__ == "__main__":
    main()
