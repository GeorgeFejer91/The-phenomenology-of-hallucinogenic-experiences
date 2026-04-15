"""Extract every excerpt from all 8 coder xlsx files into a long-format CSV.
One row per (coder, trip, item, excerpt).
Joins against taxonomy.csv to attach hierarchical item_id.
"""
import os, csv, re, warnings
os.environ["OPENBLAS_NUM_THREADS"] = "1"
warnings.filterwarnings("ignore")
import openpyxl


def slug(s):
    s = re.sub(r"[^\w\s-]", "", s).strip().lower()
    s = re.sub(r"\s+", "_", s)
    return s


def load_taxonomy(path):
    """path -> item_id lookup by path string."""
    out = {}
    with open(path, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            out[row["path"]] = row["item_id"]
    return out

FILES = [
    # (substance, block, coder, path)
    ("psilocybin", "01-10", "Alessandra", "Psilocybin_source_coding/Psilocybe_01-10_Alessandra.xlsx"),
    ("psilocybin", "01-10", "Brendan",    "Psilocybin_source_coding/Psilocybe_01-10_Brendan.xlsx"),
    ("psilocybin", "11-20", "Francesco",  "Psilocybin_source_coding/Psilocybe_11-20_Francesco.xlsx"),
    ("psilocybin", "11-20", "Susana",     "Psilocybin_source_coding/Psilocybe_11-20_Susana.xlsx"),
    ("brugmansia", "01-10", "Brendan",    "Brugmansia_source_coding/Brugmansia_01-10_Brendan.xlsx"),
    ("brugmansia", "01-10", "Noah",       "Brugmansia_source_coding/Brugmansia_01-10_Unknown_Noah.xlsx"),
    ("brugmansia", "11-20", "Alessandra", "Brugmansia_source_coding/Brugmansia_11-20_Alessandra.xlsx"),
    ("brugmansia", "11-20", "Alessio",    "Brugmansia_source_coding/Brugmansia_11-20_Alessio.xlsx"),
]

OUT = "../data/raw_codings.csv"
PROJECT_ROOT = "../.."


def norm_item(s):
    if s is None: return None
    s = str(s).replace('…', '').replace('\xa0', ' ').strip()
    return s or None


def norm_trip_id(raw, substance, block, position_idx):
    """Normalise trip ID. If raw is None, infer from substance+block+position."""
    if raw is None or str(raw).strip() in ('', '_'):
        # Block 01-10 → trips 01..10; block 11-20 → trips 11..20
        start = int(block.split('-')[0])
        num = start + position_idx
        prefix = "Psilocybe" if substance == "psilocybin" else "Brugmansia"
        return f"{prefix}_{num:02d}", True
    s = str(raw).strip().lstrip('#').strip().replace(' ', '_')
    # Normalise capitalisation: "Brugmansia_05" / "Psilocybe_05"
    # Already nearly there, just ensure underscore
    s = s.replace('__', '_')
    return s, False


def norm_excerpt(s):
    if s is None: return None
    s = str(s).replace('\xa0', ' ').strip()
    if s in ('', '_'): return None
    return s


def load_one(substance, block, coder, path, tax):
    wb = openpyxl.load_workbook(os.path.join(PROJECT_ROOT, path), data_only=True)
    ws = wb.active
    nrows, ncols = ws.max_row, ws.max_column
    # Find trip header rows
    trip_rows = []
    for r in range(1, nrows+1):
        if ws.cell(row=r, column=2).value == 'ID':
            trip_rows.append(r)
    rows_out = []
    for i, start in enumerate(trip_rows):
        end = trip_rows[i+1] - 1 if i+1 < len(trip_rows) else nrows
        raw_tid = ws.cell(row=start, column=3).value
        trip_id, inferred = norm_trip_id(raw_tid, substance, block, i)
        dose = ws.cell(row=start+1, column=3).value if ws.cell(row=start+1, column=2).value == 'Dose' else None
        dose = norm_excerpt(dose)
        # Walk items
        current_l1 = current_l2 = None
        for r in range(start+2, end+1):
            b = norm_item(ws.cell(row=r, column=2).value)
            c = norm_item(ws.cell(row=r, column=3).value)
            d = norm_item(ws.cell(row=r, column=4).value)
            e = ws.cell(row=r, column=5).value
            if b == '_': b = None
            if c == '_': c = None
            if d == '_': d = None
            if b: current_l1 = b; current_l2 = None
            if c: current_l2 = c
            # Build item_code for this row
            if d:
                item_code = f"{current_l1} | {current_l2} | {d}"
                l1, l2, l3 = current_l1, current_l2, d
            elif c:
                item_code = f"{current_l1} | {c}"
                l1, l2, l3 = current_l1, c, None
            elif b:
                item_code = b
                l1, l2, l3 = b, None, None
            else:
                continue
            # Skip judgment of reality
            if l1 == "Judgment of reality":
                continue
            # Skip if no occurrence
            if not isinstance(e, (int, float)) or e <= 0:
                continue
            # Collect excerpts (col F onward)
            excerpts = []
            for col in range(6, ncols+1):
                v = norm_excerpt(ws.cell(row=r, column=col).value)
                if v:
                    excerpts.append(v)
            item_id = tax.get(item_code, "")
            # If no excerpts, still emit one row with empty excerpt (occurrence was coded but no quote)
            if not excerpts:
                rows_out.append({
                    "substance": substance, "block": block, "coder": coder,
                    "trip_id": trip_id, "trip_id_inferred": inferred,
                    "dose_raw": dose,
                    "item_id": item_id, "item_code": item_code,
                    "level_1": l1, "level_2": l2, "level_3": l3,
                    "occurrence_count": int(e),
                    "excerpt_idx": 0, "excerpt": None,
                    "source_row": r,
                })
            else:
                for idx, ex in enumerate(excerpts):
                    rows_out.append({
                        "substance": substance, "block": block, "coder": coder,
                        "trip_id": trip_id, "trip_id_inferred": inferred,
                        "dose_raw": dose,
                        "item_id": item_id, "item_code": item_code,
                        "level_1": l1, "level_2": l2, "level_3": l3,
                        "occurrence_count": int(e),
                        "excerpt_idx": idx, "excerpt": ex,
                        "source_row": r,
                    })
    return rows_out


def main():
    here = os.path.dirname(__file__)
    tax = load_taxonomy(os.path.join(here, "../data/taxonomy.csv"))
    all_rows = []
    for sub, blk, coder, path in FILES:
        full = os.path.join(here, PROJECT_ROOT, path)
        if not os.path.exists(full):
            print(f"MISSING: {path}")
            continue
        rows = load_one(sub, blk, coder, path, tax)
        print(f"{coder} {sub} {blk}: {len(rows)} rows")
        all_rows.extend(rows)

    out = os.path.join(here, OUT)
    fieldnames = ["substance","block","coder","trip_id","trip_id_inferred","dose_raw",
                  "item_id","item_code","level_1","level_2","level_3","occurrence_count",
                  "excerpt_idx","excerpt","source_row"]
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in all_rows:
            w.writerow(r)
    print(f"wrote {out} ({len(all_rows)} rows)")


if __name__ == "__main__":
    main()
