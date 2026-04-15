"""Export every individuated scene across all 40 trips and both raters into
one big flat sheet — once as CSV, once as XLSX. Same content in both files.

Output:
  1.Recoding/data/stage1_all_individuations.csv
  1.Recoding/data/stage1_all_individuations.xlsx

One row per canonical scene (305 rows). Each row carries:
  - the unique scene_id code
  - trip/substance/coder-pair metadata
  - rater_status (which raters individuated it)
  - driver suffix (AB/FRAG/AMP/AMB/SOMA/RCL)
  - canonical span + description
  - each rater's own excerpt (rater_A_refs, rater_B_refs)
  - the Stage-1 verdict (CONVERGENT/MISS/GRANULARITY/AMBIGUITY) + flavour
  - verdict rationale
  - the literal narrative passage from canonical_span_start..end

Per-rater booleans (rater_A_individuated, rater_B_individuated) make it
trivial to pivot the sheet by rater if needed.
"""
# ======================================================================
# AI DIRECTIVE — read AI_DIRECTIVE.md at repo root
# Stage 1 of this project measures INTER-RATER CONSISTENCY ON SCENE
# INDIVIDUATION ONLY.  The atomic question is:
#   For every narrative passage, did BOTH raters individuate it as a
#   hallucinatory scene?
# Core analytical question for every only-one-rater scene:
#   MISS         - rater-compliance gap, OR
#   GRANULARITY  - lump-vs-split (resolved by data-coding rule), OR
#   AMBIGUITY-2a - under-specified Guideline edge-case, OR
#   AMBIGUITY-2b - phenomenon outside the hallucinatory-scene category.
# Do NOT resolve attribute-level disagreement (illusion vs incrusted,
# object-class, etc.) — that is Stage 2, deferred.  See AI_DIRECTIVE.md.
# ======================================================================
import os, csv

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

DATA = "../data"
TRIPS_DIR = "../data/trips"

COLS = [
    "scene_id",
    "trip_id",
    "substance",
    "block",
    "coder_A",
    "coder_B",
    "rater_status",
    "rater_A_individuated",
    "rater_B_individuated",
    "stage1_driver",
    "parent_scene_id",
    "stage1_verdict",
    "verdict_flavour",
    "verdict_source",
    "canonical_span_start",
    "canonical_span_end",
    "canonical_desc",
    "rater_A_excerpt",
    "rater_B_excerpt",
    "narrative_passage",
    "stage1_rationale",
    "verdict_rationale",
]


def load(path):
    with open(os.path.join(os.path.dirname(__file__), DATA, path), encoding="utf-8") as f:
        return list(csv.DictReader(f))


def parse_int(s):
    try:
        return int(s)
    except (ValueError, TypeError):
        return None


def derive_verdict(scene, verdict_row):
    """Return (verdict, flavour, source, rationale, parent) for the scene.

    For shared (_AB) scenes there is no row in stage1_verdicts.csv; the
    verdict is implicitly CONVERGENT.
    """
    if scene["rater_status"] == "both":
        return ("CONVERGENT", "", "auto", "Both raters individuated this passage as a hallucinatory scene.", "")
    if not verdict_row:
        return ("NEEDS_ADJUDICATION", "", "pending", "", "")
    return (
        verdict_row.get("verdict", ""),
        verdict_row.get("verdict_flavour", ""),
        verdict_row.get("verdict_source", ""),
        verdict_row.get("rationale", ""),
        verdict_row.get("parent_scene_id", "") or scene.get("parent_scene_id", ""),
    )


def main():
    here = os.path.dirname(__file__)
    trips = {t["trip_id"]: t for t in load("trips.csv")}
    scenes = load("scenes.csv")
    try:
        verdicts = {v["scene_id"]: v for v in load("stage1_verdicts.csv")}
    except FileNotFoundError:
        verdicts = {}

    # Sort: substance (brugmansia first), then trip number, then span start
    def sort_key(s):
        trip = trips.get(s["trip_id"], {})
        return (
            0 if trip.get("substance") == "brugmansia" else 1,
            int(s["trip_id"].split("_")[-1]),
            parse_int(s["canonical_span_start"]) or 0,
            s["scene_id"],
        )
    scenes.sort(key=sort_key)

    # Load narratives lazily, cached
    narrative_cache = {}

    def narrative_for(trip_id):
        if trip_id not in narrative_cache:
            p = os.path.join(here, TRIPS_DIR, f"{trip_id}.txt")
            narrative_cache[trip_id] = open(p, encoding="utf-8").read() if os.path.exists(p) else ""
        return narrative_cache[trip_id]

    rows = []
    for s in scenes:
        trip = trips.get(s["trip_id"], {})
        a = parse_int(s.get("canonical_span_start"))
        b = parse_int(s.get("canonical_span_end"))
        narrative = narrative_for(s["trip_id"])
        passage = ""
        if a is not None and b is not None and 0 <= a < len(narrative):
            passage = narrative[a:min(b, len(narrative))]

        status = s["rater_status"]
        rater_A_indiv = status in ("both", "only_A")
        rater_B_indiv = status in ("both", "only_B")

        verdict_row = verdicts.get(s["scene_id"])
        verdict, flavour, vsource, vrationale, parent_from_verdict = derive_verdict(s, verdict_row)

        rows.append({
            "scene_id": s["scene_id"],
            "trip_id": s["trip_id"],
            "substance": trip.get("substance", ""),
            "block": trip.get("block", ""),
            "coder_A": trip.get("coder_A", ""),
            "coder_B": trip.get("coder_B", ""),
            "rater_status": status,
            "rater_A_individuated": rater_A_indiv,
            "rater_B_individuated": rater_B_indiv,
            "stage1_driver": s.get("stage1_driver", "") or "",
            "parent_scene_id": parent_from_verdict or s.get("parent_scene_id", "") or "",
            "stage1_verdict": verdict,
            "verdict_flavour": flavour,
            "verdict_source": vsource,
            "canonical_span_start": a if a is not None else "",
            "canonical_span_end":   b if b is not None else "",
            "canonical_desc": s.get("canonical_desc", "") or "",
            "rater_A_excerpt": s.get("rater_A_refs", "") or "",
            "rater_B_excerpt": s.get("rater_B_refs", "") or "",
            "narrative_passage": passage,
            "stage1_rationale": s.get("stage1_rationale", "") or "",
            "verdict_rationale": vrationale,
        })

    # ---- CSV ----
    csv_path = os.path.join(here, DATA, "stage1_all_individuations.csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=COLS, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)
    print(f"wrote {csv_path}  ({len(rows)} rows)")

    # ---- XLSX ----
    if not HAS_OPENPYXL:
        print("openpyxl not installed; skipping .xlsx output")
        return

    xlsx_path = os.path.join(here, DATA, "stage1_all_individuations.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "all_individuations"

    # Header row
    ws.append(COLS)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2B2B2B")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Verdict-class colouring
    VERDICT_FILL = {
        "CONVERGENT":   PatternFill("solid", fgColor="D1FAE5"),
        "MISS":         PatternFill("solid", fgColor="FEE2E2"),
        "GRANULARITY":  PatternFill("solid", fgColor="DBEAFE"),
        "AMBIGUITY":    PatternFill("solid", fgColor="FEF3C7"),
    }
    AMB_2B_FILL = PatternFill("solid", fgColor="EDE9FE")

    for r in rows:
        ws.append([r[c] for c in COLS])
        row_idx = ws.max_row
        verdict = r["stage1_verdict"]
        fill = VERDICT_FILL.get(verdict)
        if verdict == "AMBIGUITY" and r["verdict_flavour"] == "2b":
            fill = AMB_2B_FILL
        if fill:
            # Tint just the scene_id + verdict columns
            scene_id_cell = ws.cell(row=row_idx, column=COLS.index("scene_id") + 1)
            verdict_cell = ws.cell(row=row_idx, column=COLS.index("stage1_verdict") + 1)
            scene_id_cell.fill = fill
            verdict_cell.fill = fill

    # Column widths
    widths = {
        "scene_id": 32, "trip_id": 16, "substance": 12, "block": 8,
        "coder_A": 14, "coder_B": 14,
        "rater_status": 10, "rater_A_individuated": 10, "rater_B_individuated": 10,
        "stage1_driver": 8, "parent_scene_id": 28,
        "stage1_verdict": 14, "verdict_flavour": 8, "verdict_source": 10,
        "canonical_span_start": 10, "canonical_span_end": 10,
        "canonical_desc": 55, "rater_A_excerpt": 55, "rater_B_excerpt": 55,
        "narrative_passage": 80, "stage1_rationale": 55, "verdict_rationale": 80,
    }
    for i, col in enumerate(COLS, start=1):
        letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[letter].width = widths.get(col, 18)

    ws.freeze_panes = "B2"  # scene_id column + header
    ws.auto_filter.ref = ws.dimensions

    # Data row alignment — top & wrap for long text
    wrap_al = Alignment(vertical="top", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_al

    wb.save(xlsx_path)
    print(f"wrote {xlsx_path}  ({len(rows)} rows)")

    # Summary
    from collections import Counter
    by_verdict = Counter(r["stage1_verdict"] for r in rows)
    by_substance = Counter(r["substance"] for r in rows)
    print()
    print("Verdict distribution:")
    for k, v in by_verdict.most_common():
        print(f"  {k:<20} n={v}")
    print("\nSubstance distribution:")
    for k, v in by_substance.most_common():
        print(f"  {k:<14} n={v}")


if __name__ == "__main__":
    main()
